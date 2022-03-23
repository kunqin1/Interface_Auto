# coding=utf-8
"""
操作excel的内容
"""
import sys
import os
path = os.getcwd()
# print(path)
base_path = os.path.split(path)[0]
sys.path.append(path)
sys.path.append(base_path)
import openpyxl
from base.my_logger import logger
from Util.handle_path import cases_dir


class HandExcel:

    def load_excel(self):
        # 可以传入参数，excel_path 传入地址
        # load_excel(self,excel_path)
        """
        加载excel
        """
        # open_excel = openpyxl.load_workbook(base_path + excel_path)
        try:
            open_excel = openpyxl.load_workbook(cases_dir)
            # print(base_path + "/case/用例.xlse")
            return open_excel
        except:
            logger.error("加载 {} 失败".format(cases_dir))

    def get_sheet_data(self, index=None):
        """
        加载某个sheet的全部内容
        """
        try:
            # 获取全部的sheet名称存在sheet_name对象中，（sheet_name相当于一个字典）
            sheet_name = self.load_excel().sheetnames
            # index就相当于字典的索引
            if index is None:
                index = 0
            data = self.load_excel()[sheet_name[index]]
            return data
        except:
            logger.error("获取某个excel的全部数据失败")

    def get_cell(self, row, cols):
        """
        获取某个单元格的内容的数据
        """
        data = self.get_sheet_data().cell(row=row, column=cols).value
        # 实际调用 self.get_sheet_data().cell(1,2).value
        return data

    def get_rows(self):
        """
        获取所有行数
        """
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        """
        获取某一行的内容
        row:行数
        """
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        """
        写入数据
        :param row:
        :param cols:
        :param value:
        :return:
        """
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(base_path + cases_dir)
        wb.close()

    def get_columns_value(self, key=None):
        """
        获取某一列得数据
        :param key:
        :return:
        """
        columns_list = []
        if key == None:
            key = 'A'
            columns_list_data = self.get_sheet_data()[key]
            for i in columns_list_data:
                columns_list.append(i.value)
            return columns_list

    def get_rows_number(self, case_id):
        """
        获取行号
        :param case_id:
        :return:
        """
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return num

    def get_excel_data(self):
        """
        获取excel里面所有的数据
        """
        data_list = []
        for i in range(self.get_rows() - 3):
            data_list.append(self.get_rows_value(i + 2))

        return data_list


excel_data = HandExcel()

if __name__ == '__main__':
    exec = HandExcel()
    excel_data.get_excel_data()
    # print(exec.get_rows_value(2))
